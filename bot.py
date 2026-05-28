"""
OpenWA WhatsApp Bot
A simple bot that listens to incoming messages via webhook
and auto-replies using the OpenWA REST API.
"""

import os
import re
import logging
import requests
import pyotp
from dotenv import load_dotenv
from flask import Flask, request, jsonify
from openai import OpenAI
import threading

import storage
import attendance_handlers as ah
import group_handlers as gh
from users import is_admin
import rbac
import whatsapp
import ecommerce_handlers as eh

# Load .env file (must be called before os.getenv)
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
load_dotenv(_env_path)

# ──────────────────────────────────────────────────────────────
# Configuration — update these values to match your OpenWA setup
# ──────────────────────────────────────────────────────────────
from whatsapp import send_text, OPENWA_BASE_URL, OPENWA_SESSION_ID
BOT_PORT = int(os.getenv("BOT_PORT", "5000"))
BOT_PHONE = os.getenv("BOT_PHONE", "")
# WhatsApp LID (Linked ID) — the internal ID WhatsApp uses for @mentions
# Find it in the logs: the body will show @<LID> when someone tags the bot
BOT_LID = os.getenv("BOT_LID", "")

# GitHub Models configuration (Runs natively on Azure!)
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")

# Initialize AI Client
if GITHUB_TOKEN:
    ai_client = OpenAI(
        base_url="https://models.inference.ai.azure.com",
        api_key=GITHUB_TOKEN,
    )
else:
    ai_client = None

# ──────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────
# Flask App
# ──────────────────────────────────────────────────────────────
app = Flask(__name__)

# ──────────────────────────────────────────────────────────────
# Maintenance mode
# ──────────────────────────────────────────────────────────────
_maintenance_mode = False
try:
    _maintenance_mode = (storage.get_setting("maintenance_mode", "0") == "1")
except Exception:
    pass

def is_maintenance() -> bool:
    return _maintenance_mode

def set_maintenance(enabled: bool) -> None:
    global _maintenance_mode
    _maintenance_mode = enabled
    storage.set_setting("maintenance_mode", "1" if enabled else "0")
    logger.info("🔧 Maintenance mode %s", "ENABLED" if enabled else "DISABLED")

# ──────────────────────────────────────────────────────────────
# Command handlers
# ──────────────────────────────────────────────────────────────
def cmd_set_password(chat_id: str, phone: str, raw_body: str) -> None:
    """Handle: set ambri pass <password>"""
    # Extract password from the original (non-lowered) message
    # "set ambri pass myPassword123" → "myPassword123"
    parts = raw_body.split(maxsplit=3)
    if len(parts) < 4 or not parts[3].strip():
        send_text(chat_id, "⚠️ Usage: *set ambri pass <your_password>*")
        return

    password = parts[3].strip()
    storage.set_password(phone, password)
    send_text(chat_id, "✅ Password saved and encrypted successfully!")


def cmd_set_totp(chat_id: str, phone: str, raw_body: str) -> None:
    """Handle: set ambri totp <totp_secret>"""
    # Extract TOTP secret from the original (non-lowered) message
    # "set ambri totp 7XCU6AG2..." → "7XCU6AG2..."
    parts = raw_body.split(maxsplit=3)
    if len(parts) < 4 or not parts[3].strip():
        send_text(chat_id, "⚠️ Usage: *set ambri totp <your_totp_secret>*")
        return

    totp_secret = parts[3].strip()

    # Validate the TOTP secret format
    try:
        pyotp.TOTP(totp_secret).now()
    except Exception:
        send_text(chat_id, "❌ Invalid TOTP secret. Please provide a valid base32 secret key.")
        return

    storage.set_totp_secret(phone, totp_secret)
    send_text(chat_id, "✅ TOTP secret saved and encrypted successfully!")


def cmd_generate_code(chat_id: str, phone: str) -> None:
    """Handle: generate code"""
    try:
        creds = storage.get_credentials(phone)
    except Exception:
        send_text(chat_id, "❌ Encryption Error: Data is locked with a different ENCRYPTION_KEY. Please ensure the ENCRYPTION_KEY in your server's .env matches the one on your Mac.")
        return

    if not creds:
        send_text(chat_id, "⚠️ You haven't set up yet.\nPlease set your password and TOTP secret first:\n\n• *set ambri pass <password>*\n• *set ambri totp <secret>*")
        return

    if not creds["password"]:
        send_text(chat_id, "⚠️ Password not set.\nUse: *set ambri pass <password>*")
        return

    if not creds["totp_secret"]:
        send_text(chat_id, "⚠️ TOTP secret not set.\nUse: *set ambri totp <secret>*")
        return

    # Generate the TOTP code and combine with password
    totp = pyotp.TOTP(creds["totp_secret"])
    totp_code = totp.now()
    login_code = f"{creds['password']}{totp_code}"

    send_text(chat_id, "🔑 Your login code:\n⏱️ _Expires in ~30 seconds_\n📋 _Tap & hold pesan di bawah untuk copy_")
    send_text(chat_id, login_code)


def cmd_ai(chat_id: str, prompt: str, media: dict = None, quoted_text: str = None) -> None:
    """Handle: AI chat (default fallback)"""
    if not ai_client:
        send_text(chat_id, "⚠️ AI Chat is not configured yet. Please configure GITHUB_TOKEN in the .env file.")
        return

    if not prompt.strip() and not media and not quoted_text:
        return
        
    user_content = []
    
    if quoted_text:
        user_content.append({"type": "text", "text": f"Context/Quoted message:\n\"{quoted_text}\"\n\nQuestion/Prompt:\n{prompt.strip()}"})
    elif prompt.strip():
        user_content.append({"type": "text", "text": prompt.strip()})
    else:
        if media:
            user_content.append({"type": "text", "text": "Tolong jelaskan gambar ini."})

    if media and "data" in media and "mimetype" in media:
        base64_data = media["data"]
        mime_type = media["mimetype"]
        user_content.append({
            "type": "image_url",
            "image_url": {
                "url": f"data:{mime_type};base64,{base64_data}"
            }
        })
    
    try:
        response = ai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "You are a helpful, friendly AI assistant. Keep your responses concise and well-formatted for WhatsApp."
                },
                {
                    "role": "user",
                    "content": user_content,
                }
            ],
            temperature=0.7,
            max_tokens=800,
        )
        
        reply = response.choices[0].message.content
        send_text(chat_id, reply)
    except Exception as e:
        logger.error("❌ GitHub Models AI error: %s", e)
        send_text(chat_id, f"❌ Sorry, I encountered an error while processing your request.")


def cmd_spam(chat_id: str, raw_body: str, sender_id: str, pushname: str = "") -> None:
    """Handle: spam <nomor_wa> [jumlah]"""
    parts = raw_body.split()
    if len(parts) < 2:
        send_text(chat_id, "⚠️ Usage: *spam <nomor_wa> [jumlah]*\nContoh: *spam 628123456789 3*\n_(Maksimal 5 untuk mencegah ban)_")
        return

    target = parts[1].strip()
    import re
    target = re.sub(r'\D', '', target)
    if not target:
        return send_text(chat_id, "⚠️ Nomor tidak valid.")

    target_chat_id = f"{target}@c.us"

    count = 3
    if len(parts) >= 3 and parts[2].isdigit():
        count = int(parts[2])

    # LIMIT MAX SPAM
    MAX_SPAM = 5
    if count > MAX_SPAM:
        send_text(chat_id, f"⚠️ Jumlah terlalu banyak! Dibatasi maksimal {MAX_SPAM} pesan untuk mencegah bot dibanned.")
        count = MAX_SPAM

    if count < 1:
        count = 1

    # ── Resolve caller display from bot_users registry (primary source) ──
    caller_display = None

    # Try @c.us JID first, then @lid JID
    bot_user = storage.get_bot_user(sender_id)
    if not bot_user and not sender_id.endswith("@lid"):
        # Also try by raw phone number as fallback
        raw_phone = sender_id.split("@")[0]
        bot_user = storage.get_bot_user_by_phone(raw_phone)

    if bot_user:
        name = bot_user.get("pushname") or ""
        phone_num = bot_user.get("phone") or ""
        parts_display = []
        if name:
            parts_display.append(f"*{name}*")
        if phone_num:
            parts_display.append(f"wa.me/{phone_num}")
        if parts_display:
            caller_display = " ".join(parts_display)

    # Fallback: old logic (attendance username / pushname / phone)
    if not caller_display:
        import storage as _st
        att_users = _st.get_attendance_users()
        display_name = ""
        for alias, u in att_users.items():
            if u.get("owner_chat_id") == sender_id:
                display_name = u.get("username", "")
                break

        if display_name:
            if display_name.isdigit():
                clean_num = display_name
                if clean_num.startswith("08"):
                    clean_num = "628" + clean_num[2:]
                caller_display = f"wa.me/{clean_num}"
            else:
                caller_display = f"*{display_name}*"
        elif pushname:
            caller_display = f"*{pushname}*"
        elif "@lid" in sender_id:
            caller_display = "seseorang"
        else:
            caller_display = f"wa.me/{sender_id.split('@')[0]}"

    send_text(chat_id, f"⏳ Sedang mengirim {count} pesan ke {target}...")

    def _spam_worker():
        import time, json as _json

        success_count = 0
        fail_count = 0

        for i in range(count):
            result = send_text(
                target_chat_id,
                f"🚨 *PANGGILAN URGENT!* 🚨\nKamu dipanggil oleh {caller_display}! "
                f"Tolong cek WhatsApp sekarang. ({i+1}/{count})"
            )

            if result and result.get("success") is not False:
                success_count += 1
            else:
                fail_count += 1
                # Try to detect "number not on WhatsApp" via 500 on the first attempt
                if i == 0:
                    error_raw = result.get("error", "") if result else ""
                    try:
                        err_json = _json.loads(error_raw) if isinstance(error_raw, str) else error_raw
                    except Exception:
                        err_json = {}
                    status_code = err_json.get("statusCode") or ""
                    if str(status_code) == "500":
                        send_text(
                            chat_id,
                            f"❌ Gagal mengirim ke *{target}*.\n"
                            f"Kemungkinan nomor tidak terdaftar di WhatsApp atau tidak dapat dijangkau.\n"
                            f"_(Error: {err_json.get('message', error_raw)})_"
                        )
                        return  # Abort — no point retrying

            time.sleep(2)

        # Final report
        if success_count == count:
            send_text(chat_id, f"✅ Selesai! *{success_count}/{count}* pesan terkirim ke {target}.")
        elif success_count > 0:
            send_text(chat_id, f"⚠️ Selesai dengan sebagian gagal. *{success_count}/{count}* pesan terkirim ke {target}.")
        else:
            send_text(chat_id, f"❌ Semua pesan gagal dikirim ke {target}. Periksa nomor dan coba lagi.")

    threading.Thread(target=_spam_worker).start()

def cmd_batch_download(chat_id: str) -> None:
    import openpyxl
    import io
    import base64
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Batch Message"
    ws.append(["Phone", "Message"])
    ws.append(["6283822145705", "Halo, ini pesan test batch!"])
    ws.append(["08123456789", "Pesan ini juga dikirim massal."])
    
    # Adjust column width for visibility
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 50
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    b64_data = base64.b64encode(output.read()).decode('utf-8')
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = "batch_template.xlsx"
    caption = "📄 *Template Batch Message*\n\nSilakan isi kolom Phone dan Message. Lalu balas (reply) atau lampirkan kembali file ini dengan caption:\n*batch send*"
    whatsapp.send_file(chat_id, b64_data, mimetype, filename, caption)

def cmd_batch_send(chat_id: str, media: dict) -> None:
    if not media or "data" not in media:
        send_text(chat_id, "⚠️ Harap lampirkan file Excel (.xlsx) bersamaan dengan pesan *batch send* (sebagai caption).")
        return
        
    mimetype = media.get("mimetype", "")
    # Check if the mimetype loosely looks like excel or it has a valid extension
    if "spreadsheet" not in mimetype and "excel" not in mimetype and not mimetype.endswith("xlsx"):
        # Sometimes mimetype isn't perfectly detected, so we'll log it but proceed to try parsing
        logger.warning(f"Batch send mimetype looks weird: {mimetype}, trying to parse anyway.")
        
    import base64
    import io
    import openpyxl
    import time
    import threading
    
    try:
        raw_data = base64.b64decode(media["data"])
        wb = openpyxl.load_workbook(io.BytesIO(raw_data))
        ws = wb.active
    except Exception as e:
        send_text(chat_id, f"❌ Gagal membaca file Excel. Pastikan Anda mengunggah format .xlsx yang valid.\nDetail: {e}")
        return
        
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        send_text(chat_id, "⚠️ File Excel kosong atau tidak ada data nomor (minimal 1 baris setelah header).")
        return
        
    header = rows[0]
    if str(header[0]).lower() != "phone" or str(header[1]).lower() != "message":
        send_text(chat_id, "⚠️ Format Header salah. Pastikan sel A1 = 'Phone' dan sel B1 = 'Message'.")
        return
        
    targets = []
    for row in rows[1:]:
        if not row[0]: continue
        
        phone_raw = str(row[0]).strip()
        message = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        
        import re
        phone = re.sub(r'\D', '', phone_raw)
        if phone.startswith("08"):
            phone = "628" + phone[2:]
            
        if phone and message:
            targets.append((phone, message))
            
    if not targets:
        send_text(chat_id, "⚠️ Tidak ada data valid (Nomor dan Pesan) yang bisa dikirim.")
        return
        
    send_text(chat_id, f"⏳ Memproses pengiriman ke *{len(targets)}* nomor. Jeda 2 detik per pesan untuk menghindari banned...")
    
    def _batch_worker():
        success = 0
        fail = 0
        for i, (phone, msg) in enumerate(targets):
            target_jid = f"{phone}@c.us"
            res = whatsapp.send_text(target_jid, msg)
            if res and res.get("success") is not False:
                success += 1
            else:
                fail += 1
            time.sleep(2) # 2s delay
            
        send_text(chat_id, f"✅ *Batch Broadcast Selesai!*\n\n✔️ Berhasil: {success}\n❌ Gagal: {fail}")
        
    threading.Thread(target=_batch_worker).start()


def cmd_start(chat_id: str, sender_id: str, pushname: str = "") -> None:
    """
    Handle: start  (DM only)

    Registers the sender in the bot_users table.
    - Already registered → friendly reply, no DB change.
    - New user → fetch profile pic, save to DB, send welcome card.
    """
    # Try to resolve LID to @c.us so we store the stable phone-based JID
    jid = sender_id
    if sender_id.endswith("@lid"):
        from group_handlers import resolve_lid_to_cus
        jid = resolve_lid_to_cus(sender_id)

    # Already registered?
    existing = storage.get_bot_user(jid)
    if existing:
        name = existing.get("pushname") or "kamu"
        send_text(chat_id, f"👋 Halo *{name}*! Kamu sudah terdaftar sebelumnya.")
        return

    # Extract phone number from JID
    phone = jid.split("@")[0] if "@" in jid else jid

    # Detect LID
    lid = sender_id if sender_id.endswith("@lid") else None

    # Fetch profile picture (non-blocking — returns None on failure)
    profile_pic = whatsapp.get_profile_pic(jid)

    # Save to DB
    is_new = storage.register_bot_user(
        jid=jid,
        lid=lid,
        phone=phone,
        pushname=pushname or None,
        profile_pic=profile_pic,
    )

    if is_new:
        logger.info("📋 New bot user registered: jid=%s, phone=%s, name=%s", jid, phone, pushname)
        lines = ["🎉 *Selamat Datang di Bot!*\n"]
        lines.append("✅ Kamu sudah terdaftar. Berikut data yang disimpan:\n")
        if pushname:
            lines.append(f"👤 Nama   : *{pushname}*")
        lines.append(f"📱 Nomor  : `{phone}`")
        lines.append(f"🔗 JID    : `{jid}`")
        if lid:
            lines.append(f"🔒 LID    : `{lid}`")
        if profile_pic:
            lines.append(f"🖼️ Foto   : {profile_pic}")
        lines.append("\nKetik *help* untuk melihat daftar perintah yang tersedia.")
        send_text(chat_id, "\n".join(lines))
    else:
        # Race condition — another thread registered in between
        send_text(chat_id, "👋 Kamu sudah terdaftar sebelumnya.")


# ──────────────────────────────────────────────────────────────
# Registration helper
# ──────────────────────────────────────────────────────────────

def _is_registered(sender_id: str) -> bool:
    """
    Check if a sender has registered via the 'start' command.
    Tries JID → phone → LID lookups so it works for both @c.us and @lid senders.
    """
    if storage.get_bot_user(sender_id):
        return True
    phone = sender_id.split("@")[0]
    if storage.get_bot_user_by_phone(phone):
        return True
    if sender_id.endswith("@lid") and storage.get_bot_user_by_lid(sender_id):
        return True
    return False


def cmd_list_bot_users(chat_id: str) -> None:
    """Handle: list bot users  — shows all registered bot users."""
    users = storage.get_all_bot_users()
    if not users:
        send_text(chat_id, "👥 Belum ada user yang terdaftar ke bot.")
        return

    lines = [f"👥 *Bot Users ({len(users)} terdaftar)*\n"]
    for u in users:
        name = u.get("pushname") or "-"
        phone = u.get("phone") or u.get("jid", "").split("@")[0]
        jid = u.get("jid", "-")
        lid = u.get("lid")
        pic = "✅" if u.get("profile_pic") else "❌"
        banned = "🚫 *BANNED*" if u.get("is_banned") else "🟢 Aktif"
        registered = (u.get("registered_at") or "")[:10]  # YYYY-MM-DD

        lines.append(f"👤 *{name}* | {banned}")
        lines.append(f"   • Nomor : `{phone}`")
        lines.append(f"   • JID   : `{jid}`")
        if lid:
            lines.append(f"   • LID   : `{lid}`")
        lines.append(f"   • Foto  : {pic}  | Daftar: `{registered}`")

    send_text(chat_id, "\n".join(lines))


def cmd_ban_user(chat_id: str, raw_body: str) -> None:
    """Handle: bot ban <nomor/jid>  — ban a registered bot user."""
    parts = raw_body.split(maxsplit=2)
    if len(parts) < 3:
        send_text(chat_id, "⚠️ Usage: *bot ban <nomor>*\nContoh: *bot ban 628123456789*")
        return
    identifier = parts[2].strip()
    
    if rbac.is_protected(identifier):
        send_text(chat_id, "🛡️ Tidak dapat membanned Super Admin atau Bot.")
        return
        
    if storage.ban_bot_user(identifier):
        send_text(chat_id, f"🚫 User `{identifier}` telah di-*BAN*. Mereka tidak bisa menggunakan bot lagi.")
    else:
        send_text(chat_id, f"❌ User `{identifier}` tidak ditemukan di database bot users.\nPastikan nomor sudah pernah mengirim *start* ke bot.")


def cmd_unban_user(chat_id: str, raw_body: str) -> None:
    """Handle: bot unban <nomor/jid>  — lift ban from a registered bot user."""
    parts = raw_body.split(maxsplit=2)
    if len(parts) < 3:
        send_text(chat_id, "⚠️ Usage: *bot unban <nomor>*\nContoh: *bot unban 628123456789*")
        return
    identifier = parts[2].strip()
    if storage.unban_bot_user(identifier):
        send_text(chat_id, f"✅ Ban untuk user `{identifier}` telah *dicabut*. Mereka bisa menggunakan bot kembali.")
    else:
        send_text(chat_id, f"❌ User `{identifier}` tidak ditemukan di database bot users.")


# ──────────────────────────────────────────────────────────────
# "Who made the bot?" detector
# ──────────────────────────────────────────────────────────────
_CREATOR_PATTERNS = [
    # English patterns
    r"who\s+(made|create[ds]?|built|develop(ed|s)?|wrote|own[s]?|design(ed|s)?)\s.*(bot|this)",
    r"who\s+(is|are)\s+(the\s+)?(creator|maker|developer|author|owner)\s*(of\s+)?(this\s+)?(bot)?",
    r"who.*(made|create[ds]?|built|develop(ed)?)\s*(this\s+)?(bot|you)",
    r"(bot|you).*(made|created|built|developed)\s+by\s+who",

    # Indonesian patterns
    r"siapa\s+(yang\s+)?(buat|bikin|develop|ngembangin|kembang(in|kan)?|ciptain|ciptakan|biki?n)\s*(bot|ini)",
    r"(bot|ini)\s+(di)?(buat|bikin|develop|kembang(in|kan)?)\s+(oleh\s+)?siapa",
    r"(bot|ini)\s+buatan\s+(siapa|mana)",
    r"(pembuat|pencipta|developer|creator)\s*(bot|nya)\s*(siapa|ini)",
    r"siapa\s+(pembuat|pencipta|developer|creator)\s*(bot|nya)?",
]
_CREATOR_RE = re.compile("|".join(f"({p})" for p in _CREATOR_PATTERNS), re.IGNORECASE)


def _is_asking_about_creator(text: str) -> bool:
    """Return True if the message is asking who made/created the bot."""
    return bool(_CREATOR_RE.search(text))

_SANDROID_PATTERNS = [
    r"who\s+(is|are)\s+sandroid",
    r"siapa\s+(itu\s+)?sandroid",
]
_SANDROID_RE = re.compile("|".join(f"({p})" for p in _SANDROID_PATTERNS), re.IGNORECASE)

def _is_asking_who_is_sandroid(text: str) -> bool:
    return bool(_SANDROID_RE.search(text))


# ──────────────────────────────────────────────────────────────
# Message router
# ──────────────────────────────────────────────────────────────
def handle_message(data: dict) -> None:
    """
    Process an incoming message and decide whether to reply.
    """
    # Log the full incoming data for debugging
    logger.info("📦 Raw message data: %s", data)

    # Skip messages sent by the bot itself to avoid infinite loops
    if data.get("fromMe", False):
        logger.info("⏭️ Skipping own message")
        return

    # Keep original body for extracting case-sensitive values (password, totp)
    raw_body = (data.get("body") or "").strip()
    from_id = data.get("from", "")
    is_group = data.get("isGroup", False)

    # Determine the chat to reply to
    chat_id = data.get("chatId") or from_id

    # In group chats, the actual sender is often missing in OpenWA webhooks.
    # We can extract it from the message ID: false_{group_id}_{hash}_{sender_id}
    msg_id = data.get("id", "")
    if is_group and "_" in msg_id:
        parts = msg_id.split("_")
        if len(parts) >= 4:
            sender_id = parts[3]
        else:
            sender_id = data.get("sender") or data.get("author") or from_id
    else:
        # In private chats, 'from' is the sender
        sender_id = data.get("sender") or data.get("author") or from_id
    
    # ── Mute intercept ───────────────────────────────────────
    if is_group and storage.is_muted(chat_id, sender_id):
        logger.info("🔇 Message from muted user %s in %s. Deleting.", sender_id, chat_id)
        msg_id = data.get("id")
        if msg_id:
            whatsapp.delete_message(chat_id, msg_id, for_everyone=True)
        return

    # Extract phone number (strip @c.us / @g.us suffix)
    phone = from_id.split("@")[0] if "@" in from_id else from_id

    logger.info("📩 Message from %s (sender=%s, chatId=%s): %s", from_id, sender_id, chat_id, raw_body)

    # ── Group mention filter ─────────────────────────────────
    # In group chats, only respond when the bot is @mentioned.
    # Strip the @mention tag from the message so commands parse normally.
    if is_group:
        mentioned_jids = (
            data.get("mentionedJidList")
            or data.get("mentionedJids")
            or []
        )
        bot_jid = f"{BOT_PHONE}@c.us" if BOT_PHONE else ""

        # WhatsApp uses LID (Linked ID) for @mentions in the body,
        # e.g. "@39656188063751 hello" instead of "@628xxx hello".
        # Check both BOT_LID and BOT_PHONE for compatibility.
        bot_mentioned = (
            (bot_jid and bot_jid in mentioned_jids)
            or (BOT_LID and f"@{BOT_LID}" in raw_body)
            or (BOT_PHONE and f"@{BOT_PHONE}" in raw_body)
        )

        if not bot_mentioned:
            logger.info("⏭️ Group message without bot mention, skipping")
            return

        # Strip the @mention tag from the message body so commands parse cleanly
        if BOT_LID and f"@{BOT_LID}" in raw_body:
            raw_body = raw_body.replace(f"@{BOT_LID}", "").strip()
        elif BOT_PHONE and f"@{BOT_PHONE}" in raw_body:
            raw_body = raw_body.replace(f"@{BOT_PHONE}", "").strip()
        else:
            # Fallback: strip the first @mention token
            raw_body = re.sub(r"@\S+", "", raw_body, count=1).strip()

    body = raw_body.lower()

    # ── Ban gate ──────────────────────────────────────────────────
    # This runs BEFORE everything — banned users get no other response.
    if storage.is_bot_user_banned(sender_id):
        send_text(chat_id, "🚫 Kamu telah dibanned dari bot ini. Hubungi admin jika ini adalah kesalahan.")
        return

    # ── Quoted Message Target Injection ────────────────────────
    # If the user replies to a message with a command like "user mute",
    # automatically append the sender of the quoted message as the target.
    quoted_msg = data.get("quotedMessage", {})
    if isinstance(quoted_msg, dict):
        quoted_id = quoted_msg.get("id", "")
        if "_" in quoted_id:
            parts = quoted_id.split("_")
            target_id = ""
            if len(parts) >= 4:
                target_id = parts[3]  # Group message
            elif len(parts) == 3:
                target_id = parts[1]  # Private message
                
            if target_id:
                cmds_requiring_target = [
                    "user mute", "user unmute", "user kick", 
                    "user add", "admin add", "admin remove"
                ]
                # If body is exactly the command without arguments, append target_id
                if body in cmds_requiring_target:
                    raw_body += f" {target_id}"
                    body += f" {target_id.lower()}"

    def check_rbac(feature: str) -> bool:
        if rbac.has_permission(sender_id, feature):
            return True
        send_text(chat_id, f"❌ Anda tidak memiliki akses ke fitur: *{feature}*")
        return False

    # ── 'start' command — DM only, intercept BEFORE maintenance block ──
    # So new users can always register even when maintenance is active.
    if body == "start" and not is_group:
        pushname_val = data.get("pushname") or data.get("notifyName") or ""
        cmd_start(chat_id, sender_id, pushname_val)
        return

    # ── Registration gate ─────────────────────────────────────────
    # 'hello', 'ping', 'help' are free for all. Everything else
    # requires the user to have sent 'start' first.
    _FREE_COMMANDS = {"hello", "ping", "help"}
    if body not in _FREE_COMMANDS and not _is_registered(sender_id):
        send_text(
            chat_id,
            "⚠️ Kamu belum terdaftar ke bot ini.\n\n"
            "Ketik *start* (via chat pribadi ke bot) untuk mendaftar terlebih dahulu, "
            "kemudian kamu bisa menggunakan semua fitur yang tersedia."
        )
        return

    # ── Determine if sender is admin ────────────────────────
    user_is_admin = is_admin(chat_id) or is_admin(sender_id)

    # ── Maintenance mode commands (admin only) ────────────────
    if body.startswith("maintenance on"):
        if check_rbac("maintenance"):
            set_maintenance(True)
            send_text(chat_id, "🔧 Mode maintenance AKTIF.\nSemua command dari non-admin akan diblokir.")
    elif body.startswith("maintenance off"):
        if check_rbac("maintenance"):
            set_maintenance(False)
            send_text(chat_id, "✅ Mode maintenance NONAKTIF.\nBot dapat digunakan kembali oleh semua user.")
    elif body == "maintenance status" or body == "maintenance":
        if check_rbac("maintenance"):
            status = "AKTIF 🔴" if is_maintenance() else "NONAKTIF 🟢"
            send_text(chat_id, f"🔧 Status Maintenance saat ini: *{status}*")

    # Block non-admin users when maintenance is active
    if is_maintenance() and not user_is_admin:
        send_text(chat_id, "🔧 *Bot sedang dalam maintenance*\n\nMohon maaf, bot sedang dalam perbaikan. Silakan coba lagi nanti. 🙏")
        return

    # ── Command routing ──────────────────────────────────────
    if body.startswith("set ambri pass "):
        if check_rbac("set ambri pass"):
            cmd_set_password(chat_id, phone, raw_body)

    elif body.startswith("set ambri totp "):
        if check_rbac("set ambri totp"):
            cmd_set_totp(chat_id, phone, raw_body)

    elif body == "generate code":
        if check_rbac("generate code"):
            cmd_generate_code(chat_id, phone)

    elif body == "hello":
        send_text(chat_id, "hello too 👋")

    elif body == "ping":
        send_text(chat_id, "pong 🏓")
        
    elif body == "my id":
        send_text(chat_id, f"🆔 ID Anda adalah: *{sender_id}*\n\n_(Kirim ID ini ke admin jika Anda membutuhkan akses role)_")

    elif body.startswith("spam "):
        if check_rbac("spam"):
            pushname = data.get("pushname") or data.get("notifyName") or ""
            cmd_spam(chat_id, raw_body, sender_id, pushname)
            
    elif body == "batch download":
        if check_rbac("batch download"):
            cmd_batch_download(chat_id)
            
    elif body == "batch send":
        if check_rbac("batch send"):
            cmd_batch_send(chat_id, data.get("media"))

    elif body.startswith("create store "):
        if check_rbac("store admin"): eh.cmd_store_create(chat_id, raw_body)
    elif body.startswith("store "):
        if body.startswith("store create"):
            if check_rbac("store admin"): eh.cmd_store_create(chat_id, raw_body)
        elif body == "store list":
            eh.cmd_store_list(chat_id)
        elif body.startswith("store follow"):
            eh.cmd_store_follow(chat_id, raw_body, sender_id)
        elif body.startswith("store unfollow"):
            eh.cmd_store_unfollow(chat_id, raw_body, sender_id)
        elif body.startswith("store group"):
            eh.cmd_store_group(chat_id, raw_body)
        elif body.startswith("store location"):
            eh.cmd_store_location(chat_id, raw_body)
        elif body.startswith("store cs"):
            eh.cmd_store_cs(chat_id, raw_body, sender_id)
        elif body.endswith(" open") or body.endswith(" close"):
            if check_rbac("store admin"): eh.cmd_store_open_close(chat_id, raw_body)

    elif body == "product download":
        if check_rbac("store admin"): eh.cmd_product_download(chat_id)
    elif body.startswith("product update"):
        if check_rbac("store admin"): eh.cmd_update_product(chat_id, raw_body, data.get("media"))
    elif body.startswith("product list"):
        eh.cmd_product_list(chat_id, raw_body)
    elif body.startswith("product search"):
        eh.cmd_product_search(chat_id, raw_body)
    elif body.startswith("product cart"):
        eh.cmd_product_cart(chat_id, raw_body, sender_id)
    elif body.startswith("product buy"):
        eh.cmd_product_buy(chat_id, raw_body, sender_id)

    elif body == "payment method list" or body == "payment method status":
        eh.cmd_payment_method_list(chat_id)
    elif body.startswith("payment method active"):
        if check_rbac("store admin"): eh.cmd_payment_method_active(chat_id, raw_body)
    elif body.startswith("payment method nonactive"):
        if check_rbac("store admin"): eh.cmd_payment_method_inactive(chat_id, raw_body)
    elif body.startswith("payment method set"):
        if check_rbac("store admin"): eh.cmd_payment_method_set(chat_id, raw_body)

    elif body == "order type" or body == "order type status":
        eh.cmd_order_type_list(chat_id)
    elif body.startswith("order type active"):
        if check_rbac("store admin"): eh.cmd_order_type_active(chat_id, raw_body)
    elif body.startswith("order type nonactive"):
        if check_rbac("store admin"): eh.cmd_order_type_inactive(chat_id, raw_body)

    elif body.startswith("payment list"):
        eh.cmd_payment_list(chat_id, sender_id)
    elif body.startswith("payment cancel"):
        eh.cmd_payment_cancel(chat_id, raw_body, sender_id)
    elif body.startswith("payment "):
        eh.cmd_payment(chat_id, raw_body, sender_id)

    elif body == "help":
        lines = ["🤖 *Bot Commands*\n"]

        # ── 💬 WhatsApp General ─────────────────────────────────────────
        wa_general_cmds = []
        if rbac.has_permission(sender_id, "start"): wa_general_cmds.append("• *start* — Daftar ke bot (simpan data kamu)")
        if rbac.has_permission(sender_id, "hello"): wa_general_cmds.append("• *hello* — Say hello")
        if rbac.has_permission(sender_id, "ping"):  wa_general_cmds.append("• *ping* — Check if bot is alive")
        if rbac.has_permission(sender_id, "help"):  wa_general_cmds.append("• *help* — Show this help message")
        if rbac.has_permission(sender_id, "spam"):  wa_general_cmds.append("• *spam <nomor> [jumlah]* — Spam pesan (max 5)")
        if rbac.has_permission(sender_id, "ai"):    wa_general_cmds.append("• *[Kirim pesan apa saja]* — AI akan merespon")
        
        if wa_general_cmds:
            lines.append("💬 *WhatsApp General*")
            lines.extend(wa_general_cmds)
            lines.append("")

        # ── 📢 WhatsApp Message ─────────────────────────────────────────
        wa_msg_cmds = []
        if rbac.has_permission(sender_id, "batch download"): wa_msg_cmds.append("• *batch download* — Unduh template excel untuk pesan massal")
        if rbac.has_permission(sender_id, "batch send"):     wa_msg_cmds.append("• *batch send* — Kirim pesan massal (lampirkan file excel)")
        
        if wa_msg_cmds:
            lines.append("📢 *WhatsApp Message*")
            lines.extend(wa_msg_cmds)
            lines.append("")

        # ── 📅 Attendance ──────────────────────────────────────────────
        att_cmds = []
        if rbac.has_permission(sender_id, "checkin"):      att_cmds.append("• *checkin [alias]* — Absen masuk")
        if rbac.has_permission(sender_id, "checkout"):     att_cmds.append("• *checkout [alias]* — Absen pulang")
        if rbac.has_permission(sender_id, "list history"): att_cmds.append("• *list history [alias] [week/month]* — Cek riwayat absen")
        
        if att_cmds:
            lines.append("📅 *Attendance*")
            lines.extend(att_cmds)
            lines.append("")

        # ── ⚙️ Attendance Configuration ───────────────────────────────
        att_conf_cmds = []
        if rbac.has_permission(sender_id, "set auto"):                  att_conf_cmds.append("• *set auto on/off [alias]* — Set automasi harian")
        if rbac.has_permission(sender_id, "set checkin timerange"):     att_conf_cmds.append("• *set checkin timerange [alias] HH:MM HH:MM* — Waktu acak masuk")
        if rbac.has_permission(sender_id, "set checkout timerange"):    att_conf_cmds.append("• *set checkout timerange [alias] HH:MM HH:MM* — Waktu acak pulang")
        if rbac.has_permission(sender_id, "set notes"):                 att_conf_cmds.append("• *set notes [alias] [notes]* — Custom notes absen")
        if rbac.has_permission(sender_id, "clear notes"):               att_conf_cmds.append("• *clear notes [alias]* — Reset notes absen")
        if rbac.has_permission(sender_id, "list location"):             att_conf_cmds.append("• *list location* — Lihat daftar semua lokasi")
        if rbac.has_permission(sender_id, "add location"):              att_conf_cmds.append("• *add location [nama] [lat,lng]* — Tambah lokasi baru")
        if rbac.has_permission(sender_id, "set location"):              att_conf_cmds.append("• *set location [nama] [lat,lng]* — Ubah kordinat lokasi")
        if rbac.has_permission(sender_id, "attendance list users"):     att_conf_cmds.append("• *attendance list users* — Lihat attendance user terdaftar")
        if rbac.has_permission(sender_id, "attendance add user"):       att_conf_cmds.append("• *attendance add user <alias> <user> <pass> <imei>* — Tambah attendance user")
        if rbac.has_permission(sender_id, "attendance login"):          att_conf_cmds.append("• *attendance login [alias]* — Login paksa/refresh token")
        if rbac.has_permission(sender_id, "attendance register imei"):  att_conf_cmds.append("• *attendance register imei [alias]* — Daftarkan IMEI saat ini")
        if rbac.has_permission(sender_id, "attendance generate device id"): att_conf_cmds.append("• *attendance generate device id [alias]* — Generate IMEI baru")
        
        if att_conf_cmds:
            lines.append("⚙️ *Attendance Configuration*")
            lines.extend(att_conf_cmds)
            lines.append("")

        # ── 🤖 Bot User Management ─────────────────────────────────────
        bot_user_cmds = []
        if rbac.has_permission(sender_id, "bot users"): bot_user_cmds.append("• *bot users* — Lihat semua user yang terdaftar ke bot")
        if rbac.has_permission(sender_id, "bot ban"):   bot_user_cmds.append("• *bot ban <nomor>* — Ban user (blokir akses ke bot)")
        if rbac.has_permission(sender_id, "bot unban"): bot_user_cmds.append("• *bot unban <nomor>* — Cabut ban user")
        if rbac.has_permission(sender_id, "set role"):  bot_user_cmds.append("• *set role <nomor> <role>* — Ubah role pengguna")
        
        if bot_user_cmds:
            lines.append("🤖 *Bot User Management*")
            lines.extend(bot_user_cmds)
            lines.append("")

        # ── 👥 Group Management ────────────────────────────────────────
        grp_mgmt_cmds = []
        if rbac.has_permission(sender_id, "group create"): grp_mgmt_cmds.append("• *group create <nama> [nomor1,nomor2]* — Buat grup baru")
        if rbac.has_permission(sender_id, "group update"): grp_mgmt_cmds.append("• *group update <nama> | <deskripsi>* — Update info grup")
        if rbac.has_permission(sender_id, "group users"):  grp_mgmt_cmds.append("• *group users* — List member grup")
        if rbac.has_permission(sender_id, "group leave"):  grp_mgmt_cmds.append("• *group leave* — Keluar dari grup")
        
        if grp_mgmt_cmds:
            lines.append("👥 *Group Management*")
            lines.extend(grp_mgmt_cmds)
            lines.append("")

        # ── 🛡️ Group Admin Management ──────────────────────────────────
        grp_admin_cmds = []
        if rbac.has_permission(sender_id, "admin add"):    grp_admin_cmds.append("• *admin add <nomor>* — Jadikan admin")
        if rbac.has_permission(sender_id, "admin remove"): grp_admin_cmds.append("• *admin remove <nomor>* — Hapus admin")
        if rbac.has_permission(sender_id, "user add"):     grp_admin_cmds.append("• *user add <nomor>* — Tambah member")
        if rbac.has_permission(sender_id, "user kick"):    grp_admin_cmds.append("• *user kick <nomor>* — Keluarkan member")
        if rbac.has_permission(sender_id, "user mute"):    grp_admin_cmds.append("• *user mute <nomor>* — Bisu user (pesan otomatis dihapus)")
        if rbac.has_permission(sender_id, "user unmute"):  grp_admin_cmds.append("• *user unmute <nomor>* — Batal bisu user")
        if rbac.has_permission(sender_id, "check id"):     grp_admin_cmds.append("• *check id @user* — Cek LID/nomor user")
        
        if grp_admin_cmds:
            lines.append("🛡️ *Group Admin Management*")
            lines.extend(grp_admin_cmds)
            lines.append("")

        # ── 🔑 RBAC Management ─────────────────────────────────────────
        rbac_cmds = []
        if rbac.has_permission(sender_id, "rbac list users"): rbac_cmds.append("• *rbac list users* — List users, role, dan fitur aktif")
        if rbac.has_permission(sender_id, "rbac download"):   rbac_cmds.append("• *rbac download* — Download template Excel RBAC")
        if rbac.has_permission(sender_id, "rbac upload"):     rbac_cmds.append("• *(kirim file Excel)* + *rbac upload* — Upload & Terapkan RBAC")
        
        if rbac_cmds:
            lines.append("🔑 *RBAC Management*")
            lines.extend(rbac_cmds)
            lines.append("")

        # ── 🔧 Maintenance ─────────────────────────────────────────────
        maint_cmds = []
        if rbac.has_permission(sender_id, "maintenance on"):     maint_cmds.append("• *maintenance on* — Aktifkan mode maintenance")
        if rbac.has_permission(sender_id, "maintenance off"):    maint_cmds.append("• *maintenance off* — Matikan mode maintenance")
        if rbac.has_permission(sender_id, "maintenance status"): maint_cmds.append("• *maintenance status* — Cek status maintenance")
        
        if maint_cmds:
            lines.append("🔧 *Maintenance*")
            lines.extend(maint_cmds)
            lines.append("")

        # ── 🔐 Ambri ───────────────────────────────────────────────────
        ambri_cmds = []
        if rbac.has_permission(sender_id, "set ambri pass"): ambri_cmds.append("• *set ambri pass <password>* — Set password")
        if rbac.has_permission(sender_id, "set ambri totp"): ambri_cmds.append("• *set ambri totp <secret>* — Set TOTP secret")
        if rbac.has_permission(sender_id, "generate code"):  ambri_cmds.append("• *generate code* — Generate login code")
        
        if ambri_cmds:
            lines.append("🔐 *Ambri*")
            lines.extend(ambri_cmds)
            lines.append("")

        # ── 🛍️ Ecommerce & Store ───────────────────────────────────────
        ecom_cmds = []
        ecom_admin_cmds = []

        if rbac.has_permission(sender_id, "store admin"):
            ecom_admin_cmds.append("• *create store <name> <phone> <address>* — Buat toko baru")
            store = eh.es.get_store_by_admin(sender_id)
            if store:
                ecom_admin_cmds.append(f"• *store {store['id']} open/close* — Buka/tutup toko")
                ecom_admin_cmds.append("• *product download* — Download template produk (Excel)")
                ecom_admin_cmds.append(f"• *product update {store['id']}* + Excel — Update produk toko")
                ecom_admin_cmds.append("• *payment method active/nonactive <name>* — Status metode bayar")
                ecom_admin_cmds.append("• *payment method set <name> <url>* — Set gambar QR metode bayar")
                ecom_admin_cmds.append("• *order type active/nonactive <name>* — Status tipe order")
        
        if rbac.has_permission(sender_id, "store list"): ecom_cmds.append("• *store list* — Lihat semua toko")
        if rbac.has_permission(sender_id, "store follow"): ecom_cmds.append("• *store follow/unfollow <id>* — Ikuti/Berhenti ikuti toko")
        if rbac.has_permission(sender_id, "store group"): ecom_cmds.append("• *store group <id>* — Link grup komunitas toko")
        if rbac.has_permission(sender_id, "store location"): ecom_cmds.append("• *store location <id>* — Lokasi toko")
        if rbac.has_permission(sender_id, "store cs"): ecom_cmds.append("• *store cs <id> <tanya>* — Tanya admin toko")
        
        if rbac.has_permission(sender_id, "product list"): ecom_cmds.append("• *product list <id>* — Lihat daftar produk toko")
        if rbac.has_permission(sender_id, "product search"): ecom_cmds.append("• *product search <keyword>* — Cari produk")
        if rbac.has_permission(sender_id, "product cart"): 
            ecom_cmds.append("• *product cart add/remove <id> [qty]* — Kelola keranjang")
            ecom_cmds.append("• *product cart list* — Lihat keranjang saya")
            ecom_cmds.append("• *product cart process <store_id>* — Checkout keranjang")
        if rbac.has_permission(sender_id, "product buy"): ecom_cmds.append("• *product buy <id> [qty]* — Beli produk langsung")
        if rbac.has_permission(sender_id, "payment method list"): ecom_cmds.append("• *payment method list* — Lihat metode bayar")
        if rbac.has_permission(sender_id, "order type"): ecom_cmds.append("• *order type* — Lihat tipe order")
        if rbac.has_permission(sender_id, "payment"): ecom_cmds.append("• *payment <method> <type> <cart|id_qty>* — Lakukan pembayaran")
        if rbac.has_permission(sender_id, "payment list"): ecom_cmds.append("• *payment list* — List order saya")
        if rbac.has_permission(sender_id, "payment cancel"): ecom_cmds.append("• *payment cancel <order_id>* — Batalkan order")

        if ecom_admin_cmds:
            lines.append("🛍️ *Store Admin*")
            lines.extend(ecom_admin_cmds)
            lines.append("")
            
        if ecom_cmds:
            lines.append("🛒 *Ecommerce*")
            lines.extend(ecom_cmds)
            lines.append("")

        lines.append("Made with 🤖 and ❤️\nBy Sandroid")

        help_text = "\n".join(lines)
        send_text(chat_id, help_text)


    # ── Attendance User Management routing ────────────────────────
    elif body.startswith("attendance add user "):
        if check_rbac("attendance add user"):
            ah.adduser_cmd(send_text, chat_id, raw_body)
    elif body.startswith("attendance login ") or body == "attendance login":
        if check_rbac("attendance login"):
            ah.login_cmd(send_text, chat_id, raw_body)
    elif body.startswith("attendance register imei "):
        if check_rbac("attendance register imei"):
            ah.register_imei_cmd(send_text, chat_id, raw_body)
    elif body.startswith("attendance generate device id"):
        if check_rbac("attendance generate device id"):
            ah.gendeviceid_cmd(send_text, chat_id, raw_body)
    elif body == "attendance list users":
        if check_rbac("attendance list users"):
            ah.users_cmd(send_text, chat_id)
    elif body in ("list bot users", "bot users"):
        if check_rbac("bot users"):
            cmd_list_bot_users(chat_id)
    elif body.startswith("bot ban "):
        if check_rbac("bot ban"):
            cmd_ban_user(chat_id, raw_body)
    elif body.startswith("bot unban "):
        if check_rbac("bot unban"):
            cmd_unban_user(chat_id, raw_body)

    # ── Attendance routing ────────────────────────────────────────
    elif body.startswith("checkin ") or body == "checkin":
        if check_rbac("checkin"):
            ah.masuk_cmd(send_text, chat_id, raw_body)
    elif body.startswith("checkout ") or body == "checkout":
        if check_rbac("checkout"):
            ah.pulang_cmd(send_text, chat_id, raw_body)
    elif body.startswith("list history"):
        if check_rbac("list history"):
            ah.history_cmd(send_text, chat_id, raw_body)
    elif body.startswith("set auto"):
        if check_rbac("set auto"):
            ah.auto_cmd(send_text, chat_id, raw_body)
    elif body.startswith("set checkin timerange"):
        if check_rbac("set checkin timerange"):
            ah.set_checkin_timerange_cmd(send_text, chat_id, raw_body)
    elif body.startswith("set checkout timerange"):
        if check_rbac("set checkout timerange"):
            ah.set_checkout_timerange_cmd(send_text, chat_id, raw_body)
    elif body.startswith("set notes "):
        if check_rbac("set notes"):
            ah.setnotes_cmd(send_text, chat_id, raw_body)
    elif body.startswith("clear notes "):
        if check_rbac("clear notes"):
            ah.clearnotes_cmd(send_text, chat_id, raw_body)
    elif body.startswith("set location "):
        if check_rbac("set location"):
            ah.setlocation_cmd(send_text, chat_id, raw_body)
    elif body == "list location":
        if check_rbac("list location"):
            ah.location_list_cmd(send_text, chat_id)
    elif body.startswith("add location "):
        if check_rbac("add location"):
            ah.addlocation_cmd(send_text, chat_id, raw_body)


    # ── Group & Admin routing ───────────────────────────────────
    elif body.startswith("group create "):
        if check_rbac("group create"):
            gh.group_create_cmd(send_text, chat_id, raw_body)
    elif body.startswith("group update"):
        if check_rbac("group update"):
            gh.group_update_cmd(send_text, chat_id, raw_body, data)
    elif body == "group users":
        if check_rbac("group users"):
            gh.group_users_cmd(send_text, chat_id, data)
    elif body == "group leave":
        if check_rbac("group leave"):
            gh.group_leave_cmd(send_text, chat_id, data)

    elif body.startswith("admin add "):
        if check_rbac("admin add"):
            gh.admin_add_cmd(send_text, chat_id, raw_body, data)
    elif body.startswith("admin remove "):
        if check_rbac("admin remove"):
            gh.admin_remove_cmd(send_text, chat_id, raw_body, data)
    elif body.startswith("user add "):
        if check_rbac("user add"):
            gh.user_add_cmd(send_text, chat_id, raw_body, data)
    elif body.startswith("user kick "):
        if check_rbac("user kick"):
            gh.user_kick_cmd(send_text, chat_id, raw_body, data)
    elif body.startswith("user mute "):
        if check_rbac("user mute"):
            gh.user_mute_cmd(send_text, chat_id, raw_body, data)
    elif body.startswith("user unmute "):
        if check_rbac("user unmute"):
            gh.user_unmute_cmd(send_text, chat_id, raw_body, data)
    elif body.startswith("check id"):
        if check_rbac("check id"):
            gh.check_id_cmd(send_text, chat_id, data, bot_lid=BOT_LID, bot_phone=BOT_PHONE)

    # ── RBAC routing ─────────────────────────────────────────
    elif body == "rbac list users":
        if check_rbac("rbac list users"):
            msg = rbac.list_users_with_roles()
            send_text(chat_id, msg)

    elif body == "rbac download":
        if check_rbac("rbac download"):
            send_text(chat_id, "⏳ Generating RBAC Excel template...")
            b64_data = rbac.generate_template_b64()
            whatsapp.send_file(
                chat_id=chat_id,
                base64_data=b64_data,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename="rbac_template.xlsx",
                caption="✅ Ini adalah konfigurasi RBAC saat ini. Edit dan kirim kembali file ini dengan caption `rbac upload`."
            )

    elif body == "rbac upload":
        if check_rbac("rbac upload"):
            media = data.get("media")
            if media and "data" in media:
                base64_data = media["data"]
                send_text(chat_id, "⏳ Memproses file RBAC...")
                msg = rbac.parse_template_b64(base64_data)
                send_text(chat_id, msg)
            else:
                send_text(chat_id, "❌ Tidak ada file document/excel yang dilampirkan. Pastikan Anda melampirkan file Excel saat mengetik `rbac upload`.")

    elif body.startswith("set role "):
        if check_rbac("set role"):
            # format: set role 628123 admin
            parts = body.split(maxsplit=3)
            if len(parts) >= 4:
                target_number = parts[2]
                role_name = parts[3]
                
                if rbac.is_protected(target_number):
                    send_text(chat_id, "🛡️ Tidak dapat mengubah role dari Super Admin atau Bot.")
                else:
                    msg = rbac.assign_role(sender_id, target_number, role_name)
                    send_text(chat_id, msg)
            else:
                send_text(chat_id, "⚠️ Usage: *set role <nomor> <role_name>*\nContoh: *set role 628123 admin*")

    # ── "Who made the bot?" detector ────────────────────────
    elif _is_asking_who_is_sandroid(body):
        send_text(chat_id, "kepo banget")
        
    elif _is_asking_about_creator(body):
        send_text(chat_id, "🤖 Bot ini dibuat oleh *Sandroid* ✨\n\nMade with 🤖 and ❤️")

    else:
        # Default fallback: Treat as an AI prompt
        if check_rbac("ai"):
            quoted_msg = data.get("quotedMessage", {})
            quoted_text = quoted_msg.get("body") if isinstance(quoted_msg, dict) else None
            cmd_ai(chat_id, raw_body, data.get("media"), quoted_text)


@app.route("/webhook", methods=["POST"])
def webhook():
    """
    Webhook endpoint that OpenWA calls when events occur.
    Register this URL in your OpenWA session webhook settings.
    """
    payload = request.get_json(silent=True)

    if not payload:
        return jsonify({"status": "no payload"}), 400

    event = payload.get("event", "")
    data = payload.get("data", {})

    logger.info("🔔 Webhook event: %s", event)

    if event == "message.received":
        threading.Thread(target=handle_message, args=(data,)).start()
    elif event == "session.status":
        status = data.get("status", "")
        logger.info("📡 Session status changed: %s", status)

    return jsonify({"status": "ok"}), 200


@app.route("/health", methods=["GET"])
def health():
    """Simple health-check endpoint."""
    return jsonify({"status": "running", "bot": "openwa-bot"}), 200

# ──────────────────────────────────────────────────────────────
# Public OTP APIs
# ──────────────────────────────────────────────────────────────

@app.route("/api/generate-code", methods=["POST"])
def generate_code_api():
    payload = request.get_json(silent=True)
    if not payload or "phone_number" not in payload:
        return jsonify({"success": False, "error": "Missing phone_number"}), 400
        
    phone = payload["phone_number"]
    
    # Check cooldown
    allowed, time_left = storage.can_request_otp(phone, cooldown=60)
    if not allowed:
        return jsonify({"success": False, "error": "Cooldown active", "retry_in_seconds": time_left}), 429
        
    # Generate OTP
    code = storage.generate_and_save_otp(phone, expires_in=300)
    
    # Send via WhatsApp
    chat_id = f"{phone}@c.us"
    msg = f"🔐 *Kode OTP Anda adalah: {code}*\n\nBerlaku selama 5 menit. JANGAN BERIKAN KODE INI KEPADA SIAPAPUN."
    res = whatsapp.send_text(chat_id, msg)
    
    # whatsapp.send_text returns {"success": False, ...} on failure, 
    # but on success it returns the raw API response which might not have "success": True.
    if res and res.get("success") is not False:
        return jsonify({"success": True, "message": "OTP sent successfully"}), 200
    else:
        return jsonify({"success": False, "error": "Failed to send WhatsApp message", "results": res}), 500

@app.route("/api/validate-code", methods=["POST"])
def validate_code_api():
    payload = request.get_json(silent=True)
    if not payload or "phone_number" not in payload or "code" not in payload:
        return jsonify({"success": False, "error": "Missing phone_number or code"}), 400
        
    phone = payload["phone_number"]
    code = payload["code"]
    
    is_valid = storage.verify_otp(phone, code)
    if is_valid:
        return jsonify({"success": True, "message": "OTP validated successfully"}), 200
    else:
        return jsonify({"success": False, "error": "Invalid or expired OTP"}), 400


# ──────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # Initialize the database on startup
    storage.init_db()

    logger.info("=" * 50)
    logger.info("🤖 OpenWA Bot starting...")
    logger.info("   API URL    : %s", OPENWA_BASE_URL)
    logger.info("   Session ID : %s", OPENWA_SESSION_ID)
    logger.info("   Bot Port   : %s", BOT_PORT)
    logger.info("=" * 50)
    logger.info("")
    logger.info("📌 Register this webhook URL in OpenWA:")
    logger.info("   http://<your-server-ip>:%s/webhook", BOT_PORT)
    logger.info("")

    app.run(host="0.0.0.0", port=BOT_PORT, debug=False)
