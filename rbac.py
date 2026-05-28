import base64
import io
import openpyxl
import storage
from config import ADMIN_CHAT_IDS
import logging

logger = logging.getLogger(__name__)

# Keys that existed in the OLD coarse-grained RBAC (used for migration detection)
_OLD_RBAC_KEYS = {
    "attendance", "konfigurasi", "lokasi", "user_management",
    "login_code", "rbac", "ai", "maintenance", "group_management", "admin_group"
}

# ── Per-command feature keys ──────────────────────────────────────────────
DEFAULT_FEATURES = [
    # Attendance (Absensi)
    "checkin", "checkout", "list history",
    # Attendance Configuration
    "set auto", "set checkin timerange", "set checkout timerange",
    "set notes", "clear notes", "set location", "list location", "add location",
    "attendance list users", "attendance add user", "attendance login",
    "attendance register imei", "attendance generate device id",
    "set ambri pass", "set ambri totp", "generate code",
    # Bot User Management
    "bot users", "bot ban", "bot unban", "set role",
    # Group Management
    "group create", "group update", "group users", "group leave",
    # Group Admin Management
    "admin add", "admin remove", "user add", "user kick",
    "user mute", "user unmute", "check id",
    # RBAC Management
    "rbac list users", "rbac download", "rbac upload",
    # Maintenance
    "maintenance on", "maintenance off", "maintenance status",
    # WhatsApp General
    "spam", "ai", "start", "hello", "ping", "help",
    # WhatsApp Message
    "batch download", "batch send",
]

DEFAULT_ROLES = ["super admin", "admin", "user"]

_SA = {"super admin": True,  "admin": False, "user": False}
_AA = {"super admin": True,  "admin": True,  "user": False}
_ALL = {"super admin": True, "admin": True,  "user": True}

DEFAULT_MATRIX = {
    # ── Attendance (Absensi) ──────────────────────────────────────────────
    "checkin":                        _ALL.copy(),
    "checkout":                       _ALL.copy(),
    "list history":                   _ALL.copy(),
    # ── Attendance Configuration ──────────────────────────────────────────
    "set auto":                       _ALL.copy(),
    "set checkin timerange":          _ALL.copy(),
    "set checkout timerange":         _ALL.copy(),
    "set notes":                      _ALL.copy(),
    "clear notes":                    _ALL.copy(),
    "set location":                   _AA.copy(),
    "list location":                  _ALL.copy(),
    "add location":                   _AA.copy(),
    "attendance list users":          _ALL.copy(),
    "attendance add user":            _ALL.copy(),
    "attendance login":               _ALL.copy(),
    "attendance register imei":       _ALL.copy(),
    "attendance generate device id":  _ALL.copy(),
    "set ambri pass":                 _ALL.copy(),
    "set ambri totp":                 _ALL.copy(),
    "generate code":                  _ALL.copy(),
    # ── Bot User Management ───────────────────────────────────────────────
    "bot users":                      _SA.copy(),
    "bot ban":                        _SA.copy(),
    "bot unban":                      _SA.copy(),
    "set role":                       _SA.copy(),
    # ── Group Management ──────────────────────────────────────────────────
    "group create":                   _AA.copy(),
    "group update":                   _AA.copy(),
    "group users":                    _ALL.copy(),
    "group leave":                    _SA.copy(),
    # ── Group Admin Management ────────────────────────────────────────────
    "admin add":                      _AA.copy(),
    "admin remove":                   _AA.copy(),
    "user add":                       _AA.copy(),
    "user kick":                      _AA.copy(),
    "user mute":                      _AA.copy(),
    "user unmute":                    _AA.copy(),
    "check id":                       _ALL.copy(),
    # ── RBAC Management ───────────────────────────────────────────────────
    "rbac list users":                _SA.copy(),
    "rbac download":                  _SA.copy(),
    "rbac upload":                    _SA.copy(),
    # ── Maintenance ───────────────────────────────────────────────────────
    "maintenance on":                 _SA.copy(),
    "maintenance off":                _SA.copy(),
    "maintenance status":             _ALL.copy(),
    # ── WhatsApp Message ──────────────────────────────────────────────────
    "batch download":                 _AA.copy(),
    "batch send":                     _AA.copy(),
    # ── WhatsApp General ──────────────────────────────────────────────────
    "spam":                           _AA.copy(),
    "ai":                             _ALL.copy(),
    "start":                          _ALL.copy(),
    "hello":                          _ALL.copy(),
    "ping":                           _ALL.copy(),
    "help":                           _ALL.copy(),
}


def init_default_rbac():
    """
    Initialize RBAC permissions.
    - Empty DB  → insert all new defaults.
    - Old coarse keys detected → auto-migrate to new granular defaults.
    - Already using new keys → no action.
    """
    perms = storage.get_all_rbac_permissions()
    existing = {p["feature"] for p in perms}

    if existing & _OLD_RBAC_KEYS:
        # ── Migration from old coarse-grained keys ───────────────────────
        logger.info("🔄 Migrating RBAC to per-command keys...")
        new_perms = [
            {"feature": feat, "role": role, "is_active": active}
            for feat, roles in DEFAULT_MATRIX.items()
            for role, active in roles.items()
        ]
        storage.set_rbac_permissions(new_perms)
        logger.info("✅ RBAC migrated: %d permissions set", len(new_perms))
    elif not perms:
        logger.info("Initializing default RBAC permissions...")
        new_perms = [
            {"feature": feat, "role": role, "is_active": active}
            for feat, roles in DEFAULT_MATRIX.items()
            for role, active in roles.items()
        ]
        storage.set_rbac_permissions(new_perms)

def get_user_role(chat_id: str) -> str:
    """Returns the role of the user. Defaults to 'super admin' if in ADMIN_CHAT_IDS, else 'user'."""
    # Strict matching without domain suffix
    clean_id = chat_id.split('@')[0] if chat_id else ""
    for admin_id in ADMIN_CHAT_IDS:
        if admin_id.split('@')[0] == clean_id:
            return "super admin"
    
    role = storage.get_user_role(chat_id)
    if role:
        return role
    
    return "user"

def is_protected(target_id: str) -> bool:
    """Check if the target is the bot itself or a super admin."""
    import os
    clean_target = target_id.split('@')[0] if target_id else ""
    
    # Check if target is bot
    bot_phone = os.getenv("BOT_PHONE", "").split('@')[0]
    bot_lid = os.getenv("BOT_LID", "").split('@')[0]
    
    if (bot_phone and clean_target == bot_phone) or (bot_lid and clean_target == bot_lid):
        return True
        
    # Check if target is super admin
    if get_user_role(clean_target) == "super admin":
        return True
        
    return False

def has_permission(chat_id: str, feature: str) -> bool:
    """Checks if a user has permission to access a feature."""
    role = get_user_role(chat_id)
    
    # Super admin always has access, but we still respect the matrix if defined.
    # However, if it's not defined in the matrix, super admin gets access.
    
    perms = storage.get_all_rbac_permissions()
    
    # If the matrix is totally empty (shouldn't happen), fallback to defaults
    if not perms:
        init_default_rbac()
        perms = storage.get_all_rbac_permissions()
        
    for p in perms:
        if p["feature"].lower() == feature.lower() and p["role"].lower() == role.lower():
            return p["is_active"]
            
    # Fallback if rule not found in DB
    if role == "super admin":
        return True
    return False


# ── Feature group definitions (canonical order for Excel template) ──────
_FEATURE_GROUPS = [
    ("attendance", [
        "checkin", "checkout", "list history",
    ]),
    ("attendance configuration", [
        "set auto", "set checkin timerange", "set checkout timerange",
        "set notes", "clear notes",
        "set location", "list location", "add location",
        "attendance list users", "attendance add user", "attendance login",
        "attendance register imei", "attendance generate device id",
    ]),
    ("bot user management", [
        "bot users", "bot ban", "bot unban", "set role",
    ]),
    ("group management", [
        "group create", "group update", "group users", "group leave",
    ]),
    ("group admin management", [
        "admin add", "admin remove", "user add", "user kick",
        "user mute", "user unmute", "check id",
    ]),
    ("rbac management", [
        "rbac list users", "rbac download", "rbac upload",
    ]),
    ("maintenance", [
        "maintenance on", "maintenance off", "maintenance status",
    ]),
    ("whatsapp general", [
        "spam", "ai", "start", "hello", "ping", "help",
    ]),
    ("ambri", [
        "set ambri pass", "set ambri totp", "generate code",
    ]),
]

# Friendly descriptions shown in the "description" column
_FEATURE_DESC = {
    "ai":               "AI Chat — balas pesan biasa",
    "attendance":       "Attendance — checkin / checkout / history",
    "konfigurasi":      "Attendance — set auto, timerange, notes",
    "lokasi":           "Attendance — list/add location",
    "user_management":  "Attendance — kelola user absensi",
    "login_code":       "Attendance — login code / TOTP",
    "spam":             "Spam panggilan ke nomor lain",
    "rbac":             "Bot User Management — ban/unban/set role/rbac",
    "group_management": "WhatsApp — buat/edit/keluar grup",
    "admin_group":      "WhatsApp — admin add/remove, mute, check id",
    "maintenance":      "Mode maintenance (blokir non-admin)",
}

# Non-role metadata columns the parser must skip
_SKIP_COLUMNS = {"feature", "group", "description", "category", "keterangan"}


def generate_template_b64() -> str:
    """Generates a styled Excel RBAC template grouped by feature category."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RBAC Matrix"

    perms = storage.get_all_rbac_permissions()

    # Build matrix: dict[feature][role] = is_active
    matrix: dict[str, dict[str, bool]] = {}
    all_roles_set: set[str] = set(DEFAULT_ROLES)
    for p in perms:
        matrix.setdefault(p["feature"], {})[p["role"]] = p["is_active"]
        all_roles_set.add(p["role"])

    def role_sort_key(r: str) -> int:
        return {"super admin": 0, "admin": 1, "user": 2}.get(r.lower(), 99)

    roles = sorted(all_roles_set, key=role_sort_key)

    # ── Styles ────────────────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")  # dark navy
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    GROUP_FILLS  = [
        PatternFill("solid", fgColor="E8F4FD"),  # AI       – light blue
        PatternFill("solid", fgColor="EBF5EB"),  # Attend   – light green
        PatternFill("solid", fgColor="FFF3CD"),  # Bot User – light amber
        PatternFill("solid", fgColor="F5E6FF"),  # WA       – light purple
        PatternFill("solid", fgColor="FFE8E8"),  # Access   – light red
    ]
    GROUP_HEADER_FILL  = PatternFill("solid", fgColor="2E5090")
    GROUP_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ACTIVE_FILL        = PatternFill("solid", fgColor="C6EFCE")  # green tint
    NON_FILL           = PatternFill("solid", fgColor="FFCCCC")  # red tint
    CENTER             = Alignment(horizontal="center", vertical="center")
    LEFT               = Alignment(horizontal="left",   vertical="center")
    thin_side          = Side(style="thin", color="CCCCCC")
    thin_border        = Border(left=thin_side, right=thin_side,
                                top=thin_side, bottom=thin_side)

    # ── Header row ────────────────────────────────────────────────────────
    header = ["feature", "group"] + roles
    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER if col_idx > 2 else LEFT
        cell.border = thin_border

    current_row = 2

    # ── Feature rows grouped ──────────────────────────────────────────────
    for g_idx, (group_name, group_features) in enumerate(_FEATURE_GROUPS):
        row_fill = GROUP_FILLS[g_idx % len(GROUP_FILLS)]

        # --- feature data rows ---
        for feature in group_features:
            # Fallback to DEFAULT_MATRIX if feature is not yet in the DB
            feat_matrix = matrix.get(feature, DEFAULT_MATRIX.get(feature, {}))
            
            row_vals = [feature, group_name] + [
                "active" if feat_matrix.get(role, False) else "non"
                for role in roles
            ]
            ws.append(row_vals)
            for col_idx, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                if col_idx <= 2:
                    cell.fill      = row_fill
                    cell.alignment = LEFT
                    if col_idx == 1:
                        cell.font = Font(bold=True, size=10)
                else:
                    cell.alignment = CENTER
                    cell.fill = ACTIVE_FILL if val == "active" else NON_FILL
            current_row += 1

    # ── Column widths ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22   # feature key
    ws.column_dimensions["B"].width = 30   # group name
    for i, _ in enumerate(roles, start=3):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # Freeze header row
    ws.freeze_panes = "C2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return base64.b64encode(buffer.read()).decode("utf-8")


def parse_template_b64(base64_data: str) -> str:
    """Parses an uploaded base64 Excel template and applies the new RBAC matrix."""
    try:
        binary_data = base64.b64decode(base64_data)
        buffer = io.BytesIO(binary_data)
        wb = openpyxl.load_workbook(buffer)
        ws = wb.active

        # Read headers
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) < 2:
            return "❌ Format Excel tidak valid. File kosong atau tidak memiliki baris data."

        headers = [
            str(h).strip().lower() if h is not None else ""
            for h in rows[0]
        ]

        if "feature" not in headers:
            return "❌ Format Excel tidak valid. Kolom 'feature' tidak ditemukan."

        feature_idx = headers.index("feature")

        # Role columns = everything that is NOT a known metadata column
        role_columns = [
            (i, h) for i, h in enumerate(headers)
            if h and h not in _SKIP_COLUMNS
        ]

        new_perms = []
        for row in rows[1:]:
            # Skip group-header rows (feature cell is None/empty)
            if not row or feature_idx >= len(row) or not row[feature_idx]:
                continue

            feature = str(row[feature_idx]).strip()
            if not feature:
                continue

            for col_idx, role in role_columns:
                val = (
                    str(row[col_idx]).strip().lower()
                    if col_idx < len(row) and row[col_idx] is not None
                    else "non"
                )
                is_active = val in {"active", "true", "yes", "1", "aktif"}
                new_perms.append({"feature": feature, "role": role, "is_active": is_active})

        if not new_perms:
            return "❌ Tidak ada data RBAC yang valid ditemukan di file."

        storage.set_rbac_permissions(new_perms)
        feature_count = len(set(p["feature"] for p in new_perms))
        role_count    = len(set(p["role"]    for p in new_perms))
        return (
            f"✅ Konfigurasi RBAC berhasil diupdate dari Excel!\n"
            f"📊 {feature_count} fitur × {role_count} role diproses."
        )

    except Exception as e:
        logger.error(f"Error parsing RBAC Excel: {e}")
        return f"❌ Gagal memproses file Excel: {str(e)}"


def assign_role(executor_chat_id: str, target_number: str, target_role: str) -> str:
    """Assign a role to a number, adhering to RBAC assignment rules."""
    executor_role = get_user_role(executor_chat_id)
    target_role = target_role.strip().lower()
    target_id = target_number.strip()
    
    if "@" not in target_id:
        import re
        clean_num = re.sub(r'\D', '', target_id)
        if not clean_num:
            return "❌ Nomor/ID tidak valid."
        target_chat_id = f"{clean_num}@c.us"
    else:
        target_chat_id = target_id
        
    # Validation Rules
    if target_role == "super admin":
        if executor_role != "super admin":
            return "❌ Hanya Super Admin yang dapat mengangkat Super Admin baru."
            
    # Admin can assign user or admin.
    if executor_role == "admin" and target_role not in ["user", "admin"]:
        return "❌ Admin hanya dapat mengangkat role 'user' atau 'admin'."
        
    storage.set_user_role(target_chat_id, target_role)
    return f"✅ Berhasil mengubah role {target_chat_id} menjadi *{target_role}*."

def list_users_with_roles() -> str:
    """Return a formatted string listing all users, their roles, status, and features."""
    roles_db = storage.get_rbac_user_roles()
    
    for admin_id in ADMIN_CHAT_IDS:
        if admin_id not in roles_db:
            roles_db[admin_id] = "super admin"
            
    att_users = storage.get_attendance_users()
    owner_map = {}
    for alias, u in att_users.items():
        owner = u.get("owner_chat_id")
        if owner:
            if owner not in owner_map:
                owner_map[owner] = []
            owner_map[owner].append(u)

    for owner in owner_map:
        if owner not in roles_db:
            roles_db[owner] = "user"
            
    if not roles_db:
        return "Belum ada user yang terdaftar."
        
    perms = storage.get_all_rbac_permissions()
    if not perms:
        init_default_rbac()
        perms = storage.get_all_rbac_permissions()

    # Friendly label map: internal key → display name (grouped)
    FEATURE_LABELS = {
        "checkin":                       "📅 Attendance — checkin",
        "checkout":                      "📅 Attendance — checkout",
        "list history":                  "📅 Attendance — list history",
        "set auto":                      "📅 Config — set auto",
        "set checkin timerange":         "📅 Config — set checkin timerange",
        "set checkout timerange":        "📅 Config — set checkout timerange",
        "set notes":                     "📅 Config — set notes",
        "clear notes":                   "📅 Config — clear notes",
        "set location":                  "📅 Config — set location",
        "list location":                 "📅 Config — list location",
        "add location":                  "📅 Config — add location",
        "attendance list users":         "📅 Config — attendance list users",
        "attendance add user":           "📅 Config — attendance add user",
        "attendance login":              "📅 Config — attendance login",
        "attendance register imei":      "📅 Config — attendance register imei",
        "attendance generate device id": "📅 Config — generate device id",
        "set ambri pass":                "📅 Config — set ambri pass",
        "set ambri totp":                "📅 Config — set ambri totp",
        "generate code":                 "📅 Config — generate code",
        "bot users":                     "🤖 Bot Mgmt — bot users",
        "bot ban":                       "🤖 Bot Mgmt — bot ban",
        "bot unban":                     "🤖 Bot Mgmt — bot unban",
        "set role":                      "🤖 Bot Mgmt — set role",
        "group create":                  "👥 Group — group create",
        "group update":                  "👥 Group — group update",
        "group users":                   "👥 Group — group users",
        "group leave":                   "👥 Group — group leave",
        "admin add":                     "🛡️ Admin — admin add",
        "admin remove":                  "🛡️ Admin — admin remove",
        "user add":                      "🛡️ Admin — user add",
        "user kick":                     "🛡️ Admin — user kick",
        "user mute":                     "🛡️ Admin — user mute",
        "user unmute":                   "🛡️ Admin — user unmute",
        "check id":                      "🛡️ Admin — check id",
        "rbac list users":               "🔑 RBAC — rbac list users",
        "rbac download":                 "🔑 RBAC — rbac download",
        "rbac upload":                   "🔑 RBAC — rbac upload",
        "maintenance on":                "🔧 Maintenance — on",
        "maintenance off":               "🔧 Maintenance — off",
        "maintenance status":            "🔧 Maintenance — status",
        "spam":                          "💬 General — spam",
        "ai":                            "💬 General — AI Chat",
        "start":                         "💬 General — start",
        "hello":                         "💬 General — hello",
        "ping":                          "💬 General — ping",
        "help":                          "💬 General — help",
    }

    role_features = {}
    for p in perms:
        r = p["role"].lower()
        if r not in role_features:
            role_features[r] = []
        if p["is_active"]:
            role_features[r].append(p["feature"])
            
    all_features = list(set([p["feature"] for p in perms]))
    role_features["super admin"] = all_features
    
    msg = "🛡️ *RBAC Users List*\n"
    for chat_id, role in sorted(roles_db.items()):
        status = "Inactive"
        if chat_id in ADMIN_CHAT_IDS:
            status = "Active"
        else:
            owned = owner_map.get(chat_id, [])
            if owned:
                if any(u.get("active") for u in owned):
                    status = "Active"
                else:
                    status = "Inactive"
            else:
                status = "Active (No Alias)"
                
        raw_features = role_features.get(role.lower(), [])
        feature_labels = sorted(set(FEATURE_LABELS.get(f, f) for f in raw_features))
        feature_str = f"{len(feature_labels)} akses" if feature_labels else "-"
        
        display_id = chat_id.replace("@c.us", "")
        
        msg += f"\n👤 *{display_id}*\n"
        msg += f"• Role   : `{role}`\n"
        msg += f"• Status : `{status}`\n"
        msg += f"• Fitur  : {feature_str}\n"
        
    return msg


