import re
import io
import base64
import qrcode
import openpyxl
import logging
import uuid
import datetime
from datetime import timezone
import threading
import time

import whatsapp
import rbac
import ecommerce_storage as es
import storage
logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────
def parse_id_name(text: str) -> str:
    parts = text.split(maxsplit=2)
    if len(parts) >= 3:
        return parts[2].strip()
    return ""

def generate_qr_base64(data: str) -> str:
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode("utf-8")

# ──────────────────────────────────────────────────────────────
# Store Management
# ──────────────────────────────────────────────────────────────
def cmd_store_create(chat_id: str, raw_body: str) -> None:
    # E.g. create store Sandroid Shop 628222 Jl. Sudirman
    prefix = "store create " if raw_body.startswith("store create") else "create store "
    content = raw_body[len(prefix):].strip()
    
    match = re.search(r'^(.*?)\s+(\d{10,15})\s+(.*)$', content)
    if not match:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *create store <store_name> <admin_phone> <address>*\nExample: *create store Sandroid Shop 628123456789 Jl. Sudirman No 1*")
        
    store_name = match.group(1).strip()
    admin_phone = match.group(2).strip()
    address = match.group(3).strip()
    
    store_id = f"STORE-{admin_phone[-4:]}"
    
    # Auto-create group
    group_url = ""
    group_resp = whatsapp.create_group(f"Komunitas {store_name}", [f"{admin_phone}@c.us"])
    if group_resp and isinstance(group_resp, dict) and "id" in group_resp:
        group_id = group_resp["id"]
        group_url = whatsapp.get_group_invite_code(group_id)
        
    es.create_store(store_id, store_name, admin_phone, admin_phone, address, group_url=group_url)
    
    msg = f"✅ Store created!\nID: {store_id}\nName: {store_name}\nAdmin: {admin_phone}\nAddress: {address}"
    if group_url:
        msg += f"\n\n🔗 Community Group created:\n{group_url}"
    else:
        msg += "\n\n⚠️ Failed to create community group automatically."
        
    whatsapp.send_text(chat_id, msg)

def cmd_store_list(chat_id: str) -> None:
    stores = es.get_all_stores()
    if not stores:
        whatsapp.send_text(chat_id, "🛒 No stores available.")
        return
    
    lines = ["🛒 *Available Stores*"]
    for s in stores:
        status = "🟢 OPEN" if s['is_open'] else "🔴 CLOSED"
        lines.append(f"• *{s['name']}* ({s['id']}) - {status}\n  📍 {s['address']}")
    whatsapp.send_text(chat_id, "\n\n".join(lines))

def cmd_store_follow(chat_id: str, raw_body: str, sender_id: str) -> None:
    store_identifier = parse_id_name(raw_body)
    if not store_identifier:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *store follow <store_id/name>*")
        
    store = es.get_store(store_identifier)
    if not store:
        return whatsapp.send_text(chat_id, "❌ Store not found.")
        
    es.follow_store(store['id'], sender_id)
    whatsapp.send_text(chat_id, f"✅ You are now following *{store['name']}*!\nYou will receive updates via DM when new products are available.\n\nTo stop receiving updates, send: *store unfollow {store['id']}*")

def cmd_store_unfollow(chat_id: str, raw_body: str, sender_id: str) -> None:
    store_identifier = parse_id_name(raw_body)
    if not store_identifier:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *store unfollow <store_id/name>*")
        
    store = es.get_store(store_identifier)
    if not store:
        return whatsapp.send_text(chat_id, "❌ Store not found.")
        
    es.unfollow_store(store['id'], sender_id)
    whatsapp.send_text(chat_id, f"✅ You have unfollowed *{store['name']}*. You will no longer receive DM updates.")

def cmd_store_group(chat_id: str, raw_body: str) -> None:
    store_identifier = parse_id_name(raw_body)
    if not store_identifier:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *store group <store_id/name>*")
        
    store = es.get_store(store_identifier)
    if not store:
        return whatsapp.send_text(chat_id, "❌ Store not found.")
        
    if not store['group_url']:
        return whatsapp.send_text(chat_id, f"ℹ️ *{store['name']}* does not have a group link yet.")
        
    whatsapp.send_text(chat_id, f"🔗 Join the *{store['name']}* community here:\n{store['group_url']}")

def cmd_store_location(chat_id: str, raw_body: str) -> None:
    store_identifier = parse_id_name(raw_body)
    if not store_identifier:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *store location <store_id/name>*")
        
    store = es.get_store(store_identifier)
    if not store:
        return whatsapp.send_text(chat_id, "❌ Store not found.")
        
    whatsapp.send_text(chat_id, f"📍 *{store['name']}* Location:\n{store['address']}")

def cmd_store_cs(chat_id: str, raw_body: str, sender_id: str) -> None:
    parts = raw_body.split(maxsplit=3)
    if len(parts) < 4:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *store cs <store_id/name> <question>*")
        
    store_identifier = parts[2].strip()
    question = parts[3].strip()
    
    store = es.get_store(store_identifier)
    if not store:
        return whatsapp.send_text(chat_id, "❌ Store not found.")
        
    admin_phone = store['admin_phone']
    admin_jid = f"{admin_phone}@c.us"
    
    # Notify Admin
    whatsapp.send_text(admin_jid, f"🔔 *Customer Service Request*\n\nFrom: wa.me/{sender_id.split('@')[0]}\nQuestion: {question}")
    whatsapp.send_text(chat_id, "✅ Your question has been forwarded to the store admin. They will reply to you shortly.")

def cmd_store_open_close(chat_id: str, raw_body: str) -> None:
    parts = raw_body.split(maxsplit=2)
    if len(parts) < 3:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *store <store_id/name> open/close*")
        
    store_identifier = parts[1].strip()
    action = parts[2].lower()
    
    is_open = action == "open"
    if not is_open and action != "close":
        return whatsapp.send_text(chat_id, "⚠️ Please specify 'open' or 'close'.")
        
    store = es.get_store(store_identifier)
    if not store:
        return whatsapp.send_text(chat_id, "❌ Store not found.")
        
    clean_sender = sender_id.split('@')[0]
    if store['admin_phone'] != clean_sender and store['phone_number'] != clean_sender and rbac.get_user_role(sender_id) != "super admin":
        return whatsapp.send_text(chat_id, "❌ Unauthorized. You do not own this store.")
        
    es.set_store_open(store['id'], is_open)
    status_text = "🟢 OPEN" if is_open else "🔴 CLOSED"
    whatsapp.send_text(chat_id, f"✅ Store *{store['name']}* is now {status_text}.")

# ──────────────────────────────────────────────────────────────
# Product Management
# ──────────────────────────────────────────────────────────────
def cmd_product_download(chat_id: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(["id", "product name", "description", "image (url based)", "price", "stock"])
    ws.append(["PROD-001", "Bengbeng", "snack chocolate", "https://example.com/bengbeng.jpg", 3000, 10])
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    b64_data = base64.b64encode(output.read()).decode('utf-8')
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    filename = "product_template.xlsx"
    caption = "📄 *Product Template*\n\nFill this out and reply with:\n*product update <store_id>*"
    whatsapp.send_file(chat_id, b64_data, mimetype, filename, caption)

def cmd_update_product(chat_id: str, raw_body: str, media: dict) -> None:
    store_identifier = parse_id_name(raw_body)
    if not store_identifier:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *product update <store_id>* (attach the Excel file)")
        
    store = es.get_store(store_identifier)
    if not store:
        return whatsapp.send_text(chat_id, "❌ Store not found.")
        
    clean_sender = chat_id.split('@')[0]
    if store['admin_phone'] != clean_sender and store['phone_number'] != clean_sender and rbac.get_user_role(chat_id) != "super admin":
        return whatsapp.send_text(chat_id, "❌ Unauthorized. You do not own this store.")
        
    if store['is_open']:
        return whatsapp.send_text(chat_id, "❌ Store must be CLOSED before updating products to prevent transaction issues. Use *store <id> close* first.")
        
    if not media or "data" not in media:
        return whatsapp.send_text(chat_id, "⚠️ Please attach the filled Excel file.")
        
    try:
        raw_data = base64.b64decode(media["data"])
        wb = openpyxl.load_workbook(io.BytesIO(raw_data))
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return whatsapp.send_text(chat_id, "⚠️ Excel file is empty.")
            
        products = []
        for r in rows[1:]:
            if not r[0] or not r[1]: continue
            products.append({
                'id': str(r[0]),
                'name': str(r[1]),
                'description': str(r[2]) if r[2] else "",
                'image_url': str(r[3]) if r[3] else "",
                'price': float(r[4]) if r[4] else 0,
                'stock': int(r[5]) if r[5] else 0
            })
            
        es.replace_store_products(store['id'], products)
        whatsapp.send_text(chat_id, f"✅ Successfully updated {len(products)} products for *{store['name']}*!")
        
        # Notify followers
        followers = es.get_store_followers(store['id'])
        if followers:
            def _notify():
                msg = f"🔔 *New Products from {store['name']}!*\n\nCheck them out using *product list {store['id']}*\n\nTo stop updates, send: *store unfollow {store['id']}*"
                for f in followers:
                    whatsapp.send_text(f, msg)
                    time.sleep(1)
            threading.Thread(target=_notify).start()
            
    except Exception as e:
        logger.error(f"Excel Parse Error: {e}")
        whatsapp.send_text(chat_id, f"❌ Failed to parse Excel: {e}")

def cmd_product_list(chat_id: str, raw_body: str) -> None:
    store_identifier = parse_id_name(raw_body)
    products = es.get_products(store_identifier)
    
    if not products:
        return whatsapp.send_text(chat_id, "📦 No products found.")
        
    lines = [f"📦 *Product List*"]
    for p in products:
        lines.append(f"• *{p['name']}* ({p['id']})\n  💰 Rp{p['price']} | 📦 Stock: {p['stock']}")
    
    whatsapp.send_text(chat_id, "\n\n".join(lines))

def cmd_product_search(chat_id: str, raw_body: str) -> None:
    keyword = parse_id_name(raw_body)
    if not keyword:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *product search <keyword>*")
        
    products = es.search_products(keyword)
    if not products:
        return whatsapp.send_text(chat_id, f"🔍 No products found matching '{keyword}'.")
        
    lines = [f"🔍 *Search Results for '{keyword}'*"]
    for p in products:
        lines.append(f"• *{p['name']}* ({p['id']})\n  💰 Rp{p['price']} | 📦 Stock: {p['stock']}")
        
    whatsapp.send_text(chat_id, "\n\n".join(lines))

def send_welcome_products(chat_id: str, store: dict) -> None:
    whatsapp.send_text(chat_id, f"👋 Welcome to *{store['name']}* community!\n\nHere are our latest products:")
    products = es.get_products(store['id'])
    if not products:
        whatsapp.send_text(chat_id, "📦 Stay tuned for our upcoming products!")
        return
        
    lines = [f"📦 *Product List - {store['name']}*"]
    # Limit to top 10 to avoid spamming
    for p in products[:10]:
        lines.append(f"• *{p['name']}* ({p['id']})\n  💰 Rp{p['price']} | 📦 Stock: {p['stock']}")
        
    lines.append(f"\nTo buy, use: *product buy <product_id>*")
    whatsapp.send_text(chat_id, "\n\n".join(lines))

# ──────────────────────────────────────────────────────────────
# Cart & Checkout
# ──────────────────────────────────────────────────────────────
def cmd_product_cart(chat_id: str, raw_body: str, sender_id: str) -> None:
    parts = raw_body.split()
    if len(parts) < 3:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *product cart <add/remove/clear/process> ...*")
        
    action = parts[2].lower()
    
    if action == "add":
        if len(parts) < 4:
            return whatsapp.send_text(chat_id, "⚠️ Usage: *product cart add <product_id> [qty]*")
        prod_id = parts[3]
        qty = int(parts[4]) if len(parts) > 4 and parts[4].isdigit() else 1
        
        prod = es.get_product(prod_id)
        if not prod:
            return whatsapp.send_text(chat_id, "❌ Product not found.")
            
        if prod['stock'] < qty:
            return whatsapp.send_text(chat_id, f"❌ Not enough stock! Only {prod['stock']} available.")
            
        store = es.get_store(prod['store_id'])
        if not store or not store['is_open']:
            return whatsapp.send_text(chat_id, "❌ Store is currently closed.")
            
        es.add_to_cart(sender_id, prod['store_id'], prod['id'], qty)
        whatsapp.send_text(chat_id, f"🛒 Added {qty}x *{prod['name']}* to cart.")
        
    elif action == "remove":
        if len(parts) < 4:
            return whatsapp.send_text(chat_id, "⚠️ Usage: *product cart remove <product_id> [qty]*")
        prod_id = parts[3]
        qty = int(parts[4]) if len(parts) > 4 and parts[4].isdigit() else None
        es.remove_from_cart(sender_id, prod_id, qty)
        whatsapp.send_text(chat_id, f"🛒 Removed from cart.")
        
    elif action == "clear":
        es.clear_cart(sender_id)
        whatsapp.send_text(chat_id, "🛒 Cart cleared.")
        
    elif action == "process":
        if len(parts) < 4:
            return whatsapp.send_text(chat_id, "⚠️ Usage: *product cart process <store_id>*")
        store_id = parts[3]
        cart_items = es.get_cart(sender_id, store_id)
        
        if not cart_items:
            return whatsapp.send_text(chat_id, "🛒 Your cart for this store is empty.")
            
        total = 0
        lines = [f"🛒 *Checkout Summary ({store_id})*"]
        for item in cart_items:
            subtotal = item['qty'] * item['price']
            total += subtotal
            lines.append(f"• {item['qty']}x {item['name']} = Rp{subtotal}")
            
        lines.append(f"\nTotal: *Rp{total}*")
        lines.append(f"\nTo complete payment, use:\n*payment <payment_method> <order_type> <cart>*")
        whatsapp.send_text(chat_id, "\n".join(lines))
        
    elif action == "list":
        cart_items = es.get_cart(sender_id)
        if not cart_items:
            return whatsapp.send_text(chat_id, "🛒 Your cart is empty.")
            
        lines = ["🛒 *Your Cart*"]
        for item in cart_items:
            lines.append(f"• {item['qty']}x {item['name']} (Store: {item['store_id']}) = Rp{item['qty'] * item['price']}")
        whatsapp.send_text(chat_id, "\n".join(lines))

def cmd_product_buy(chat_id: str, raw_body: str, sender_id: str) -> None:
    parts = raw_body.split()
    if len(parts) < 3:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *product buy <product_id> [qty]*")
        
    prod_id = parts[2]
    qty = int(parts[3]) if len(parts) > 3 and parts[3].isdigit() else 1
    
    prod = es.get_product(prod_id)
    if not prod:
        return whatsapp.send_text(chat_id, "❌ Product not found.")
        
    store = es.get_store(prod['store_id'])
    if not store or not store['is_open']:
        return whatsapp.send_text(chat_id, "❌ Store is currently closed.")
        
    if prod['stock'] < qty:
        return whatsapp.send_text(chat_id, f"❌ Not enough stock! Only {prod['stock']} available.")
        
    total = prod['price'] * qty
    lines = [
        f"🛍️ *Direct Buy Summary*",
        f"• {qty}x {prod['name']} = Rp{total}",
        f"\nTotal: *Rp{total}*",
        f"\nTo complete payment, use:\n*payment <payment_method> <order_type> {prod['id']}_{qty}*"
    ]
    whatsapp.send_text(chat_id, "\n".join(lines))

# ──────────────────────────────────────────────────────────────
# Order Types & Payments Config
# ──────────────────────────────────────────────────────────────
def cmd_order_type_list(chat_id: str) -> None:
    types = es.get_order_types(active_only=False)
    if not types:
        return whatsapp.send_text(chat_id, "ℹ️ No order types configured.")
    lines = ["🚚 *Order Types*"]
    for t in types:
        status = "✅ Active" if t['is_active'] else "❌ Inactive"
        lines.append(f"• {t['name']} - {status}")
    whatsapp.send_text(chat_id, "\n".join(lines))

def cmd_order_type_active(chat_id: str, raw_body: str) -> None:
    name = parse_id_name(raw_body)
    if not name:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *order type active <name>*")
    es.set_order_type_status(name, True)
    whatsapp.send_text(chat_id, f"✅ Order type '{name}' is now active.")

def cmd_order_type_inactive(chat_id: str, raw_body: str) -> None:
    name = parse_id_name(raw_body)
    if not name:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *order type nonactive <name>*")
    es.set_order_type_status(name, False)
    whatsapp.send_text(chat_id, f"✅ Order type '{name}' is now inactive.")

def cmd_payment_method_list(chat_id: str) -> None:
    methods = es.get_payment_methods(active_only=False)
    if not methods:
        return whatsapp.send_text(chat_id, "ℹ️ No payment methods configured.")
    lines = ["💳 *Payment Methods*"]
    for m in methods:
        status = "✅ Active" if m['is_active'] else "❌ Inactive"
        lines.append(f"• {m['name']} - {status}")
    whatsapp.send_text(chat_id, "\n".join(lines))

def cmd_payment_method_active(chat_id: str, raw_body: str) -> None:
    name = parse_id_name(raw_body)
    if not name:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *payment method active <name>*")
    es.set_payment_method_status(name, True)
    whatsapp.send_text(chat_id, f"✅ Payment method '{name}' is now active.")

def cmd_payment_method_inactive(chat_id: str, raw_body: str) -> None:
    name = parse_id_name(raw_body)
    if not name:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *payment method nonactive <name>*")
    es.set_payment_method_status(name, False)
    whatsapp.send_text(chat_id, f"✅ Payment method '{name}' is now inactive.")

def cmd_payment_method_set(chat_id: str, raw_body: str) -> None:
    parts = raw_body.split(maxsplit=3)
    if len(parts) < 4:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *payment method set <name> <image_url>*")
    name = parts[2]
    url = parts[3]
    es.set_payment_method_image(name, url)
    whatsapp.send_text(chat_id, f"✅ Image set for '{name}'.")

# ──────────────────────────────────────────────────────────────
# Checkout & Payment
# ──────────────────────────────────────────────────────────────
def cmd_payment(chat_id: str, raw_body: str, sender_id: str) -> None:
    # payment <payment_method> <order_type> <order_id/cart>
    parts = raw_body.split()
    if len(parts) < 4:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *payment <method> <type> <cart|product_id_qty>*")
        
    method_name = parts[1]
    type_name = parts[2]
    target = parts[3]
    
    pm = es.get_payment_method(method_name)
    if not pm or not pm['is_active']:
        return whatsapp.send_text(chat_id, f"❌ Payment method '{method_name}' is not available.")
        
    order_type = next((t for t in es.get_order_types() if t['name'].lower() == type_name.lower()), None)
    if not order_type:
        return whatsapp.send_text(chat_id, f"❌ Order type '{type_name}' is not available.")
        
    items = []
    store_id = None
    
    if target == "cart":
        # Need to know which store's cart to process.
        # Check if user passed store_id as 4th arg
        if len(parts) > 4:
            store_id = parts[4]
            cart = es.get_cart(sender_id, store_id)
        else:
            # If no store passed, try to fetch all, but they must belong to 1 store for an order.
            cart = es.get_cart(sender_id)
            if not cart:
                return whatsapp.send_text(chat_id, "🛒 Your cart is empty.")
            
            stores = set(c['store_id'] for c in cart)
            if len(stores) > 1:
                return whatsapp.send_text(chat_id, "❌ You have items from multiple stores in your cart. Please process per store: *payment <method> <type> cart <store_id>*")
            store_id = cart[0]['store_id']
            
        if not cart:
            return whatsapp.send_text(chat_id, "🛒 Cart is empty.")
            
        for c in cart:
            if c['stock'] < c['qty']:
                return whatsapp.send_text(chat_id, f"❌ Stock issue: {c['name']} only has {c['stock']} left.")
            items.append({'product_id': c['product_id'], 'qty': c['qty'], 'price': c['price']})
            
        # Clean cart
        es.clear_cart(sender_id, store_id)
        
    else:
        # direct buy: target = prodID_qty
        if "_" in target:
            prod_id, qty_str = target.split("_", 1)
            qty = int(qty_str)
        else:
            prod_id = target
            qty = 1
            
        prod = es.get_product(prod_id)
        if not prod:
            return whatsapp.send_text(chat_id, "❌ Product not found.")
        if prod['stock'] < qty:
            return whatsapp.send_text(chat_id, f"❌ Not enough stock for {prod['name']}.")
            
        store_id = prod['store_id']
        items.append({'product_id': prod['id'], 'qty': qty, 'price': prod['price']})
        
    # Create Order
    order_id = f"ORD-{uuid.uuid4().hex[:6].upper()}"
    order_data = es.create_order(order_id, sender_id, store_id, order_type['name'], pm['name'], items)
    
    expires_at = datetime.datetime.fromisoformat(order_data['expires_at']).replace(tzinfo=timezone.utc).astimezone()
    expires_str = expires_at.strftime("%Y-%m-%d %H:%M:%S")
    
    caption = f"🧾 *ORDER CREATED: {order_id}*\n\n💰 Total: Rp{order_data['total_amount']}\n🚚 Type: {order_type['name']}\n💳 Method: {pm['name']}\n\n⏳ Please pay before:\n*{expires_str}*\n\nIf unpaid, it will be automatically cancelled."
    
    if pm['image_url']:
        # Download image and send
        try:
            import requests
            resp = requests.get(pm['image_url'])
            b64_img = base64.b64encode(resp.content).decode('utf-8')
            whatsapp.send_file(chat_id, b64_img, "image/jpeg", "qr.jpg", caption)
        except Exception:
            whatsapp.send_text(chat_id, caption + "\n\n❌ Failed to load payment image.")
    else:
        # Generate QR Dev
        qr_string = pm.get('dev_qr_string') or f"DEV-PAY-{order_id}-{order_data['total_amount']}"
        b64_qr = generate_qr_base64(qr_string)
        whatsapp.send_file(chat_id, b64_qr, "image/png", "qr.png", caption)
        
def cmd_payment_list(chat_id: str, sender_id: str) -> None:
    orders = es.get_orders(sender_id)
    if not orders:
        return whatsapp.send_text(chat_id, "ℹ️ You have no orders.")
        
    lines = ["🧾 *Your Orders*"]
    # Show top 5
    for o in orders[:5]:
        lines.append(f"• *{o['id']}* - {o['status'].upper()}\n  💰 Rp{o['total_amount']} | {o['created_at'][:10]}")
    
    whatsapp.send_text(chat_id, "\n\n".join(lines))

def cmd_payment_cancel(chat_id: str, raw_body: str, sender_id: str) -> None:
    parts = raw_body.split(maxsplit=2)
    if len(parts) < 3:
        return whatsapp.send_text(chat_id, "⚠️ Usage: *payment cancel <order_id>*")
        
    order_id = parts[2].strip()
    order = es.get_order(order_id)
    
    if not order:
        return whatsapp.send_text(chat_id, "❌ Order not found.")
        
    if order['user_id'] != sender_id and rbac.get_user_role(sender_id) != "super admin":
        return whatsapp.send_text(chat_id, "❌ Unauthorized.")
        
    if order['status'] != 'pending':
        return whatsapp.send_text(chat_id, f"❌ Cannot cancel order that is {order['status']}.")
        
    es.update_order_status(order_id, 'cancel')
    whatsapp.send_text(chat_id, f"✅ Order {order_id} cancelled.")
